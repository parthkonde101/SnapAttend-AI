"""Student roster Excel import ("Extending the attendance system" spec,
Part 3 — Student Import System).

Reads an uploaded `.xlsx` workbook's first sheet and upserts `Student` rows
for one panel, matched by PRN. An imported row *is* a SnapAttend account,
full stop — no separate staging table. A brand-new PRN gets a fresh
`Student` row with the administrator-issued default password
(`DEFAULT_STUDENT_PASSWORD`, hashed) and `password_changed = False`, which
forces the student through the mandatory Change Password screen on first
login (see `app/schemas/student.StudentChangePasswordRequest` and
`POST /students/me/change-password`).

Interpretation of the spec's import-behavior bullets (documented here
since the spec leaves some of this implicit):
  * "Insert new records" — a PRN not already in `students` becomes a new
    row: `panel_id` set to the panel being imported into, password set to
    the hashed default, `password_changed = False`, `is_active = True`.
  * "Update existing records" — a PRN already in `students` (in *any*
    panel) has its name/roll_number/batch/panel_id overwritten with this
    row's values and is (re-)marked `is_active = True` — the row's data and
    the panel it's imported into are both treated as the current source of
    truth. The PRN itself is never rewritten by this path (it's the match
    key, never a target field). Deliberately, the *password* is never
    touched on an update — a student who already signed in and possibly
    changed their password should not have it silently reset just because
    an admin re-uploaded the same roster (that's what the admin's explicit
    "Reset Password" action, `AdminStudentService.reset_to_default_password`,
    is for).
  * "No duplicates" — a PRN that appears more than once *within this same
    uploaded file*; only the first occurrence is processed, every later
    repeat of that PRN in the same file is counted as "Skipped." (An
    existing DB row for a PRN is always "Updated," never "Skipped" — skip
    is reserved for in-file repeats and entirely blank rows.)
  * "Validate every row" — a row missing PRN or Full Name is counted as an
    "Error" (with the 1-indexed row number and a human-readable reason),
    not silently dropped and not applied to the database.
  * "Roll Number is unique within the panel" (spec Part 13 — Panel Import
    Validation): a row whose Roll Number is already claimed by a
    *different* PRN — either earlier in this same file, or by an existing
    student already on this panel's roster — is rejected as an "Error"
    with a clear message, not silently overwritten and not silently
    skipped. A blank Roll Number is exempt (not every roster necessarily
    has one on file); uniqueness is only enforced among rows that actually
    supply one.
  * "PRN is globally unique" is enforced by construction: `Student.prn`
    carries a database-level UNIQUE constraint, and every row here is
    matched/upserted by PRN, never inserted blind — there is no code path
    in this function that could produce two `Student` rows with the same PRN.
  * A completely blank row (every expected cell empty) is neither an error
    nor a skip — it's treated as spreadsheet padding and ignored, so it
    doesn't inflate any of the four reported counts.
"""
from __future__ import annotations

import re
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.student import Student
from app.schemas.student import ExcelImportSummary, ImportRowError

# The administrator-issued starting password for every newly imported
# student, per the spec's explicit "Default password for every imported
# student: Test@123" — note the capital T. Never used to overwrite an
# existing account's password on re-import — see this module's docstring's
# "Update existing records" bullet. Always hashed (bcrypt, via
# app.core.security.hash_password) before it ever touches the database —
# never stored or logged in plaintext.
DEFAULT_STUDENT_PASSWORD = "Test@123"  # noqa: S105 - a known, documented default, always hashed before storage

# Header cells are matched case-insensitively, trimmed of surrounding
# whitespace, with a couple of common synonyms accepted per column — real
# spreadsheets are inconsistent about capitalization and phrasing far more
# often than they're inconsistent about the underlying data.
_EXPECTED_HEADERS = {
    "prn": "prn",
    "roll number": "roll_number",
    "roll no": "roll_number",
    "roll no.": "roll_number",
    "rollnumber": "roll_number",
    "name": "name",
    "full name": "name",
    "student name": "name",
    "batch": "batch",
}


class InvalidWorkbookError(Exception):
    """Raised when the uploaded file isn't a readable .xlsx, or its header
    row doesn't contain a Roll Number, a PRN, and a Full Name column."""


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


# Bugfix (carried forward from the prior roster-import work): a float
# artifact like "1032232198.0" — see `normalize_prn` below.
_FLOAT_ARTIFACT_RE = re.compile(r"^\d+\.0+$")
# Scientific notation Excel/openpyxl sometimes renders very large "numbers" in.
_SCI_NOTATION_RE = re.compile(r"^\d+(?:\.\d+)?[eE][+-]?\d+$")


def normalize_prn(value: object) -> str:
    """Convert any Excel PRN/Roll Number cell representation into a clean
    digit string.

    Excel (and, by extension, openpyxl) treats an all-digit column as
    numeric unless the column is explicitly formatted as text, so the raw
    cell value openpyxl hands back is often a Python `float`
    (e.g. `1032232198.0`) rather than the string printed on the ID card.
    Running that straight through `_cell_text` (plain `str(value).strip()`)
    would bake the trailing ".0" directly into the stored value — which
    would then never match what a student actually types. This function is
    the fix, applied to both the PRN and Roll Number columns.
    """
    if value is None:
        return ""

    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).strip()

    if isinstance(value, int):
        return str(value)

    text = str(value).strip()
    if not text:
        return ""

    if _SCI_NOTATION_RE.match(text):
        try:
            as_float = float(text)
        except ValueError:
            return text
        return str(int(as_float)) if as_float.is_integer() else text

    if _FLOAT_ARTIFACT_RE.match(text):
        return text.split(".", 1)[0]

    return text


def import_students_excel(db: Session, *, panel_id: int, file_bytes: bytes) -> ExcelImportSummary:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several distinct exception types for a corrupt/non-xlsx file
        raise InvalidWorkbookError("Could not read this file as an Excel (.xlsx) workbook.") from exc

    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise InvalidWorkbookError("The uploaded workbook is empty.") from None

    column_index: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        key = _EXPECTED_HEADERS.get(_cell_text(cell).lower())
        if key is not None and key not in column_index:
            column_index[key] = index

    if "prn" not in column_index or "name" not in column_index:
        raise InvalidWorkbookError(
            "The workbook must have 'PRN' and 'Full Name' column headers ('Roll Number' and 'Batch' are also "
            "expected, but optional)."
        )

    imported = 0
    updated = 0
    skipped = 0
    errors: list[ImportRowError] = []
    seen_prns: set[str] = set()

    # Roll Number uniqueness within this panel (spec Part 13). Two maps:
    # roll numbers already claimed by another PRN earlier in *this file*,
    # and roll numbers already on file for *this panel* in the database
    # (queried once, up front — not per-row). Both are keyed by the
    # normalized (trimmed, lowercased) roll number and store the normalized
    # PRN that currently owns it, so a row updating its own existing PRN is
    # never flagged as colliding with itself.
    seen_roll_numbers: dict[str, str] = {}
    existing_panel_roll_numbers: dict[str, str] = {
        roll.strip().lower(): prn.strip().lower()
        for prn, roll in db.execute(
            select(Student.prn, Student.roll_number).where(Student.panel_id == panel_id)
        )
        if roll and roll.strip()
    }

    def cell(row: tuple, key: str) -> str:
        index = column_index.get(key)
        if index is None or index >= len(row):
            return ""
        return _cell_text(row[index])

    def raw_cell(row: tuple, key: str) -> object:
        index = column_index.get(key)
        if index is None or index >= len(row):
            return None
        return row[index]

    default_password_hash = hash_password(DEFAULT_STUDENT_PASSWORD)

    for row_number, row in enumerate(rows, start=2):
        if row is None or all(_cell_text(value) == "" for value in row):
            continue  # entirely blank spacer row — not counted at all

        # Bugfix: PRN/Roll Number are read from the *raw* cell value (via
        # normalize_prn) rather than through `cell()`'s plain `_cell_text`/
        # `str()` — see `normalize_prn`'s docstring for why. Every other
        # column is unaffected and still goes through `cell()` as before.
        prn = normalize_prn(raw_cell(row, "prn"))
        roll_number = normalize_prn(raw_cell(row, "roll_number")) or None
        name = cell(row, "name")
        batch = cell(row, "batch") or None

        if not prn:
            errors.append(ImportRowError(row_number=row_number, message="Missing PRN."))
            continue
        if not name:
            errors.append(ImportRowError(row_number=row_number, message="Missing Full Name."))
            continue

        normalized_prn = prn.lower()
        if normalized_prn in seen_prns:
            skipped += 1
            continue
        seen_prns.add(normalized_prn)

        if roll_number:
            normalized_roll = roll_number.strip().lower()
            claimed_by_this_file = seen_roll_numbers.get(normalized_roll)
            claimed_in_panel = existing_panel_roll_numbers.get(normalized_roll)
            if claimed_by_this_file is not None and claimed_by_this_file != normalized_prn:
                errors.append(
                    ImportRowError(
                        row_number=row_number,
                        message=f"Roll Number '{roll_number}' is already used by another student earlier in this file.",
                    )
                )
                continue
            if claimed_in_panel is not None and claimed_in_panel != normalized_prn:
                errors.append(
                    ImportRowError(
                        row_number=row_number,
                        message=f"Roll Number '{roll_number}' is already used by another student in this panel.",
                    )
                )
                continue
            seen_roll_numbers[normalized_roll] = normalized_prn

        existing = db.scalar(select(Student).where(func.lower(Student.prn) == normalized_prn))
        if existing is None:
            db.add(
                Student(
                    prn=prn,
                    full_name=name,
                    roll_number=roll_number,
                    batch=batch,
                    panel_id=panel_id,
                    password_hash=default_password_hash,
                    password_changed=False,
                    is_active=True,
                )
            )
            imported += 1
        else:
            # Deliberately does not touch password_hash / password_changed —
            # see this module's docstring's "Update existing records" bullet.
            existing.full_name = name
            existing.roll_number = roll_number
            existing.batch = batch
            existing.panel_id = panel_id
            existing.is_active = True
            db.add(existing)
            updated += 1

    db.commit()

    return ExcelImportSummary(imported=imported, updated=updated, skipped=skipped, errors=errors)
