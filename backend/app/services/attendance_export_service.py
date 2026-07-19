"""Builds the teacher-facing Excel (.xlsx) attendance export.

Deliberately a separate service from `AttendanceReviewService` — that one
owns the *live, editable* review page (roster + evidence + overrides); this
one only turns an already-finalized session into a static downloadable
file. Reuses `AttendanceReviewService.build_session_review` for the roster
so the export can never drift from what the teacher actually saw and
overrode on the review page — same status field, same source of truth —
but this module then deliberately drops every field except roll number,
PRN, name, and final Present/Absent status before it ever reaches a cell.

Per spec Part 14 ("Simplified Attendance Export"), the export contains
*only* what faculty need for an official attendance record — Roll Number,
PRN, Student Name, Attendance — sorted by Roll Number. No internal IDs, no
timestamps, no panel/course names, no image paths, no AI confidence, no
marker detection, no diagnostics, no verification metadata: a single
clean, printable sheet, nothing else. (An earlier version of this export
also emitted a "Session Summary" sheet with the session's internal ID and
start/end timestamps — that sheet was removed entirely to satisfy this
spec's explicit "no internal IDs, timestamps" requirement, not just
trimmed.)
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.sorting import roll_number_sort_key
from app.models.attendance_session import AttendanceSession
from app.services.attendance_review_service import AttendanceReviewService

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_HEADER_BORDER = Border(bottom=Side(style="thin", color="9CA3AF"))
_CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")

_MIN_COLUMN_WIDTH = 10
_MAX_COLUMN_WIDTH = 48
_COLUMN_PADDING = 3


@dataclass(frozen=True)
class AttendanceExport:
    filename: str
    content: bytes


def _sanitize_filename_part(value: str) -> str:
    """Filesystem/HTTP-header-safe: session_code is already plain
    uppercase alphanumeric (see AttendanceSessionService's code alphabet),
    but this strips anything else defensively rather than assuming."""
    return re.sub(r"[^A-Za-z0-9_-]", "", value) or "Session"


def _style_header_row(ws: Worksheet, num_columns: int) -> None:
    for col_index in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _HEADER_BORDER


def _autosize_columns(ws: Worksheet) -> None:
    """openpyxl has no built-in autosize — column width is derived from
    the longest rendered cell value in that column, the same approach
    every hand-rolled openpyxl autosize implementation uses."""
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            column_letter = get_column_letter(cell.column)
            widths[column_letter] = max(widths.get(column_letter, 0), len(str(cell.value)))
    for column_letter, content_width in widths.items():
        ws.column_dimensions[column_letter].width = min(
            _MAX_COLUMN_WIDTH, max(_MIN_COLUMN_WIDTH, content_width + _COLUMN_PADDING)
        )


def _build_attendance_sheet(wb: Workbook, review) -> None:
    """Column set and ordering per spec Part 14 ("Simplified Attendance
    Export"): Roll Number, PRN, Student Name, Attendance — nothing else —
    sorted by Roll Number ascending, numeric-aware, so faculty receive the
    sheet in the exact official classroom order.
    """
    ws = wb.active
    ws.title = "Attendance"

    ws.append(["Roll Number", "PRN", "Student Name", "Attendance"])
    _style_header_row(ws, num_columns=4)
    ws.freeze_panes = "A2"

    ordered_students = sorted(review.students, key=lambda item: roll_number_sort_key(item.roll_number))
    for student in ordered_students:
        ws.append(
            [
                student.roll_number or "",
                student.prn,
                student.full_name,
                "Present" if student.status == "present" else "Absent",
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _CELL_ALIGNMENT

    _autosize_columns(ws)


def build_attendance_export(db, session: AttendanceSession) -> AttendanceExport:
    """Generate the final attendance export for one ended session.

    `db` is a live SQLAlchemy session, used only to rebuild the same
    roster the review page shows (via `AttendanceReviewService`) —
    guaranteeing the export can never disagree with what a teacher already
    reviewed and overrode. Raises nothing session-lifecycle-related itself;
    the caller (the export endpoint) is responsible for rejecting a request
    against a still-active session before this is ever called.
    """
    review = AttendanceReviewService(db).build_session_review(session)

    wb = Workbook()
    _build_attendance_sheet(wb, review)

    buffer = BytesIO()
    wb.save(buffer)

    session_name = _sanitize_filename_part(session.session_code)
    date_str = session.created_at.strftime("%Y-%m-%d")
    filename = f"Attendance_{session_name}_{date_str}.xlsx"

    return AttendanceExport(filename=filename, content=buffer.getvalue())
